#!/usr/bin/env python3
"""
EXCEL FILES SPECIALIZED REINDEXING SCRIPT
Handles .xlsx and .xls files with optimized extraction methods
Part of the parallel reindexing system for maximum performance
"""

import os
import sys
import time
import signal
import logging
import re
from datetime import datetime
from typing import Optional

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv
from azure.storage.blob import BlobServiceClient
from azure.search.documents import SearchClient
from azure.core.credentials import AzureKeyCredential
import pandas as pd
import openpyxl
import xlrd

# Load environment variables
load_dotenv()

# Configure logging to suppress pandas warnings
logging.getLogger('azure').setLevel(logging.WARNING)
logging.getLogger('urllib3').setLevel(logging.WARNING)
logging.getLogger('openpyxl').setLevel(logging.WARNING)
logging.getLogger('pandas').setLevel(logging.WARNING)

class TimeoutError(Exception):
    """Custom timeout exception"""
    pass

def timeout_handler(signum, frame):
    """Signal handler for timeouts"""
    raise TimeoutError("Excel processing timeout")

def fast_extract_excel_content(file_path: str) -> Optional[str]:
    """
    Ultra-fast Excel content extraction with multiple methods and timeout
    Returns None if extraction fails or times out
    """
    # Set 20-second timeout for Excel files
    signal.signal(signal.SIGALRM, timeout_handler)
    signal.alarm(20)
    
    try:
        content_parts = []
        
        # Method 1: Try openpyxl for .xlsx files (fastest for modern Excel)
        if file_path.lower().endswith('.xlsx'):
            try:
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                for sheet_name in workbook.sheetnames[:5]:  # Limit to first 5 sheets
                    sheet = workbook[sheet_name]
                    sheet_content = []
                    
                    # Extract first 100 rows max per sheet
                    for row_num, row in enumerate(sheet.iter_rows(max_row=100, values_only=True)):
                        if row_num > 100:
                            break
                        row_text = ' '.join(str(cell) for cell in row if cell is not None and str(cell).strip())
                        if row_text:
                            sheet_content.append(row_text)
                    
                    if sheet_content:
                        content_parts.append(f"Sheet: {sheet_name}\n" + '\n'.join(sheet_content))
                
                workbook.close()
                
                if content_parts:
                    return '\n\n'.join(content_parts)[:8000]  # Limit to 8KB
                    
            except Exception as e:
                print(f"   ‚ö†Ô∏è  openpyxl failed: {str(e)[:100]}")
        
        # Method 2: Try pandas for both .xlsx and .xls (good general purpose)
        try:
            # Read all sheets (limit to first 5)
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names[:5]
            
            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100)  # Limit rows
                    
                    # Convert to string representation
                    sheet_text = f"Sheet: {sheet_name}\n"
                    
                    # Add column headers
                    if not df.empty:
                        headers = ' | '.join(str(col) for col in df.columns)
                        sheet_text += f"Headers: {headers}\n"
                        
                        # Add data rows
                        for _, row in df.iterrows():
                            row_text = ' | '.join(str(val) for val in row.values if pd.notna(val))
                            if row_text.strip():
                                sheet_text += row_text + '\n'
                    
                    if len(sheet_text) > 50:  # Only add if we got meaningful content
                        content_parts.append(sheet_text)
                        
                except Exception as e:
                    print(f"   ‚ö†Ô∏è  pandas sheet '{sheet_name}' failed: {str(e)[:50]}")
                    continue
            
            if content_parts:
                return '\n\n'.join(content_parts)[:8000]  # Limit to 8KB
                
        except Exception as e:
            print(f"   ‚ö†Ô∏è  pandas failed: {str(e)[:100]}")
        
        # Method 3: Try xlrd for legacy .xls files
        if file_path.lower().endswith('.xls'):
            try:
                workbook = xlrd.open_workbook(file_path)
                for sheet_idx in range(min(5, workbook.nsheets)):  # First 5 sheets
                    sheet = workbook.sheet_by_index(sheet_idx)
                    sheet_name = sheet.name
                    sheet_content = []
                    
                    # Extract first 100 rows
                    for row_idx in range(min(100, sheet.nrows)):
                        row_values = []
                        for col_idx in range(sheet.ncols):
                            try:
                                cell_value = sheet.cell_value(row_idx, col_idx)
                                if cell_value:
                                    row_values.append(str(cell_value))
                            except:
                                continue
                        
                        if row_values:
                            sheet_content.append(' | '.join(row_values))
                    
                    if sheet_content:
                        content_parts.append(f"Sheet: {sheet_name}\n" + '\n'.join(sheet_content))
                
                if content_parts:
                    return '\n\n'.join(content_parts)[:8000]  # Limit to 8KB
                    
            except Exception as e:
                print(f"   ‚ö†Ô∏è  xlrd failed: {str(e)[:100]}")
        
        return None
        
    except TimeoutError:
        print(f"   ‚è∞ Timeout extracting Excel content")
        return None
        
    except Exception as e:
        print(f"   üí• Unexpected error: {str(e)[:100]}")
        return None
        
    finally:
        signal.alarm(0)  # Cancel the alarm

def needs_reprocessing(doc: dict) -> bool:
    """
    Determine if a document needs reprocessing based on content quality
    """
    content = doc.get('content', '')
    
    # Always reprocess if no content
    if not content or len(content.strip()) < 10:
        return True
    
    # Check for signs of poor extraction
    if (
        'unable to extract' in content.lower() or
        'could not parse' in content.lower() or
        'extraction failed' in content.lower() or
        'no text found' in content.lower() or
        len(content.strip()) < 50  # Very minimal content
    ):
        return True
    
    return False

def main():
    """Main execution function"""
    start_time = time.time()
    
    # Check environment variables
    required_vars = [
        'AZURE_STORAGE_CONNECTION_STRING',
        'AZURE_SEARCH_SERVICE_ENDPOINT',
        'AZURE_SEARCH_API_KEY',
        'AZURE_SEARCH_INDEX_NAME'
    ]
    
    missing_vars = [var for var in required_vars if not os.getenv(var)]
    if missing_vars:
        print(f"üí• Missing environment variables: {', '.join(missing_vars)}")
        return 1
    
    print("üìä EXCEL FILES REINDEXING STARTED")
    print("="*50)
    print(f"üïê Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    try:
        # Initialize Azure clients
        print("üîß Initializing Azure clients...")
        
        blob_service_client = BlobServiceClient.from_connection_string(
            os.getenv('AZURE_STORAGE_CONNECTION_STRING')
        )
        
        search_client = SearchClient(
            endpoint=os.getenv('AZURE_SEARCH_SERVICE_ENDPOINT'),
            index_name=os.getenv('AZURE_SEARCH_INDEX_NAME'),
            credential=AzureKeyCredential(os.getenv('AZURE_SEARCH_API_KEY'))
        )
        
        container_name = os.getenv("AZURE_STORAGE_CONTAINER_NAME", "dtce-documents")
        container_client = blob_service_client.get_container_client(container_name)
        
        print("‚úÖ Azure clients initialized!")
        print()
        
        # Get all Excel files from blob storage
        print("üîç Finding Excel files...")
        excel_files = []
        
        for blob in container_client.list_blobs():
            if blob.name.lower().endswith(('.xlsx', '.xls')):
                # Skip temporary/backup files
                if any(skip in blob.name.lower() for skip in [
                    '~$', '.tmp', 'temp', 'backup', '.bak', 
                    'recycle', 'deleted', '.keep', '.gitkeep'
                ]):
                    continue
                
                excel_files.append(blob)
        
        print(f"üìä Found {len(excel_files)} Excel files to process")
        print()
        
        if not excel_files:
            print("‚ÑπÔ∏è  No Excel files found to process")
            return 0
        
        # Process each Excel file
        processed = 0
        updated = 0
        errors = 0
        
        for i, blob in enumerate(excel_files):
            try:
                print(f"üìä [{i+1}/{len(excel_files)}] Processing: {blob.name}")
                
                # Check if document exists in search index
                try:
                    search_results = list(search_client.search(
                        search_text="*",
                        filter=f"blob_name eq '{blob.name}'",
                        select="id,content,blob_name"
                    ))
                    
                    existing_doc = search_results[0] if search_results else None
                    
                    # Skip if document exists and doesn't need reprocessing
                    if existing_doc and not needs_reprocessing(existing_doc):
                        print(f"   ‚úÖ Already processed with good content - skipping")
                        continue
                        
                except Exception as e:
                    print(f"   ‚ö†Ô∏è  Search check failed: {str(e)[:50]}")
                    existing_doc = None
                
                # Download and process the Excel file
                print(f"   üì• Downloading...")
                blob_client = container_client.get_blob_client(blob.name)
                blob_data = blob_client.download_blob().readall()
                
                # Save temporarily for processing
                temp_file = f"/tmp/temp_excel_{int(time.time())}_{os.path.basename(blob.name)}"
                with open(temp_file, 'wb') as f:
                    f.write(blob_data)
                
                try:
                    # Extract content
                    print(f"   üìä Extracting Excel content...")
                    content = fast_extract_excel_content(temp_file)
                    
                    if not content or len(content.strip()) < 20:
                        print(f"   ‚ö†Ô∏è  No meaningful content extracted")
                        errors += 1
                        continue
                    
                    # Prepare document for indexing (match schema from other scripts)
                    filename = os.path.basename(blob.name)
                    folder_path = '/'.join(blob.name.split('/')[:-1]) if '/' in blob.name else ''
                    
                    # Extract project info
                    project_name = ""
                    year = None  # Use None instead of empty string for integer fields
                    if blob.name.startswith('Projects/'):
                        parts = blob.name.split('/')
                        if len(parts) >= 3:
                            project_name = parts[2]  # e.g., "225200"
                            try:
                                if project_name.isdigit() and len(project_name) >= 3:
                                    year = int(project_name[:3]) + 2000  # Return actual integer
                            except:
                                year = None
                    
                    # Create document ID
                    document_id = blob.name.replace('/', '_').replace(' ', '_').replace('#', '').replace('.', '_')
                    document_id = re.sub(r'[^\w\-_]', '_', document_id)
                    document_id = re.sub(r'_+', '_', document_id).strip('_')
                    
                    document = {
                        "id": document_id,
                        "blob_name": blob.name,
                        "blob_url": blob_client.url,
                        "filename": filename,
                        "content_type": "application/vnd.openxml",
                        "folder": folder_path,
                        "size": blob.size or 0,
                        "content": content,
                        "last_modified": blob.last_modified.isoformat(),
                        "created_date": blob.creation_time.isoformat() if blob.creation_time else blob.last_modified.isoformat(),
                        "project_name": project_name,
                        "year": year
                    }
                    
                    # Upload to search index
                    print(f"   üì§ Uploading to search index...")
                    search_client.upload_documents([document])
                    
                    print(f"   ‚úÖ Successfully processed ({len(content)} chars)")
                    updated += 1
                    
                except Exception as e:
                    print(f"   üí• Processing failed: {str(e)[:100]}")
                    errors += 1
                
                finally:
                    # Clean up temp file
                    try:
                        os.unlink(temp_file)
                    except:
                        pass
                
                processed += 1
                
                # Progress indicator
                if processed % 10 == 0:
                    elapsed = time.time() - start_time
                    rate = processed / elapsed * 60
                    print(f"   üìä Progress: {processed}/{len(excel_files)} files ({rate:.1f}/min)")
                    print()
                
            except KeyboardInterrupt:
                print("\n‚èπÔ∏è  Interrupted by user")
                break
                
            except Exception as e:
                print(f"   üí• Unexpected error: {str(e)[:100]}")
                errors += 1
                continue
        
        # Final summary
        elapsed = time.time() - start_time
        print()
        print("üìä EXCEL FILES REINDEXING COMPLETED")
        print("="*50)
        print(f"üìä Total files processed: {processed}")
        print(f"‚úÖ Successfully updated: {updated}")
        print(f"üí• Errors: {errors}")
        print(f"‚è±Ô∏è  Total time: {elapsed/60:.1f} minutes")
        print(f"üìà Average rate: {processed/elapsed*60:.1f} files/minute")
        print(f"üïê Completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        return 0
        
    except KeyboardInterrupt:
        print("\n‚èπÔ∏è  Interrupted by user")
        return 1
        
    except Exception as e:
        print(f"\nüí• Fatal error: {e}")
        return 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
